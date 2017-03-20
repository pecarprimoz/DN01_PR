from datetime import datetime
from numpy import *
from collections import defaultdict


set_printoptions(suppress=True)

data = loadtxt("../data/ratings.csv",delimiter=",",skiprows=1)

userID=[int(i) for i in data[:,0]]
movieID=[int(i) for i in data[:,1]]
ratingID=[float(i) for i in data[:,2]]
timestamp=[datetime.fromtimestamp(i).strftime('%Y-%m-%d')
           for i in data[:,3]]
relevantData=zip(movieID,ratingID)
#vse skupne ocene, št. pojavitev
allMovies=defaultdict(list)
for mID,rating in relevantData:
    if(mID not in allMovies):
        allMovies[mID].append([0.0,0])
    allMovies[mID][0][1]+=1
    allMovies[mID][0][0]+=rating
#skupni rating (allMovies[key][0][1])
#št. pojavitev (allMovies[key][0][0])

avgRatings=defaultdict(float)
for key,value in allMovies.items():
    if(allMovies[key][0][1]!=0):
        avgRatings[key]=allMovies[key][0][0]/allMovies[key][0][1]
    else:
        avgRatings[key]=-1

topFilmi=sorted(avgRatings.items(), key=lambda v: v[1], reverse=True)
#print(topFilmi)

#težava ker je film ocenjen enkrat ali manjkrat, popravimo tako da vzamemo vse filme ki so bili ocenjeni več kot 15x
trueAvgRatings=defaultdict(float)
for key,value in allMovies.items():
    if (allMovies[key][0][1] != 0 and allMovies[key][0][1]>15):
        trueAvgRatings[key] = allMovies[key][0][0] / allMovies[key][0][1]
    else:
        trueAvgRatings[key] = -1

topTrueFilmi = sorted(trueAvgRatings.items(), key=lambda v: v[1], reverse=True)[:500]
#print(topTrueFilmi)

#uporablam openpyxl od kle naprej ker ima numpy probleme
from openpyxl import load_workbook

movieNames= load_workbook(filename="../data/moviesRMK_V1.xlsx")
useNames = movieNames['movies']
useNamesID=[]
useNamesName=[]
for i in range(2,9127):
    wholeColumn=useNames['A'+str(i)].value.split(",")
    useNamesID.append(wholeColumn[0])
    useNamesName.append(wholeColumn[1])

useNamesID=[int(i) for i in useNamesID]
skupajImena=zip(useNamesID,useNamesName)
finalShape=defaultdict(list)
for idFilma,avgRating in topTrueFilmi:
    finalShape[idFilma]=[avgRating,str()]



for skval in topTrueFilmi:
    for id,movname in zip(useNamesID,useNamesName):
        if(id==skval[0]):
            finalShape[id][1]=movname
            break
filmActors=defaultdict(list)
movieNames= load_workbook(filename="../data/castRMK_V1.xlsx")
useNames = movieNames['cast']
for i in range(2,9127):
    wholeColumn=useNames['A'+str(i)].value.split(",")
    for actor in wholeColumn[1].split("|"):
        if actor!='':
            filmActors[int(wholeColumn[0])].append(actor)

popularActors=defaultdict(int)
for key,value in finalShape.items():
    for filmID, arr in filmActors.items():
        if(key==filmID):
            for actor in arr:
                popularActors[actor]+=1
#print(popularActors)
import operator
sortedActors = sorted(popularActors.items(), key=operator.itemgetter(1))
print(sortedActors)
# print(sortedGenres)
print("10 najpopularnejših igralcev, gledamo 500 najbolje ocenjenih filmov razvrščeni v padajočem vrstnem redu.")
print("IGRALEC\t\t\t\t\t\t\tST. POJAVITEV")
print("-------------------------------------------------")
izrisActor=[]
izrisCount=[]
actorCount=0
for value in reversed(sortedActors):
    izrisActor.append(value[0])
    izrisCount.append(value[1])
    print("%-25s | %20d" % (str(value[0]), value[1]))
    actorCount+=1
    if actorCount>=10:
        break

import matplotlib.pyplot as plt
import numpy as np

y_axis = np.arange(1, len(izrisActor) + 1, 1)
plt.barh(y_axis, izrisCount[::-1], align='center')
plt.yticks(y_axis, izrisActor[::-1])
plt.show()


    #TIL, če zippaš dva arraya prej in ju uporabš v for zanki, se noče restartat, če nardiš vsakič na novo je vredu?
#key je id filma, value je array [rating,ime filma]
#for key,value in finalShape.items():
#    print(str(key)+"\t"+str(value[1]))

