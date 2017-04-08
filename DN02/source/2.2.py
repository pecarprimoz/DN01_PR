from datetime import datetime

from numpy import *
from collections import defaultdict

from openpyxl.utils import get_column_letter

set_printoptions(suppress=True)

data = loadtxt("../data/ratings.csv",delimiter=",",skiprows=1)

userID=[int(i) for i in data[:,0]]
movieID=[int(i) for i in data[:,1]]
ratingID=[float(i) for i in data[:,2]]
timestamp=[datetime.fromtimestamp(i).strftime('%Y-%m-%d')
           for i in data[:,3]]
relevantData=zip(userID,movieID,ratingID)
temp=set(userID)
#print(len(list(temp)))
#vse skupne ocene, št. pojavitev
allMovies=defaultdict(list)
for uID,mID,rating in relevantData:
    if len(allMovies[mID])!=2:
        allMovies[mID].append([])
        allMovies[mID].append([])
        allMovies[mID][0]=["?" for i in range(0,len(list(temp))+1)]
    allMovies[mID][1]=[0]

ikd=zip(userID,movieID,ratingID)
for id,mv,rt in ikd:
    tem1p= allMovies[mv]
    allMovies[mv][0][id]=rt
    allMovies[mv][1][0]+=1



topTrueFilmi = sorted(allMovies.items(), key=lambda v: v[1][1][0], reverse=True)[:100]
movNames=defaultdict(list)
from openpyxl import load_workbook
movieNames= load_workbook(filename="../data/moviesRMK_V1.xlsx")
useNames = movieNames['movies']
useNamesID=[]
useNamesName=[]
for i in range(2,9127):
    wholeColumn=useNames['A'+str(i)].value.split(",")
    useNamesID.append(wholeColumn[0])
    useNamesName.append(wholeColumn[1])

useNamesID=[int(i) for i in useNamesID]
skupajImena=zip(useNamesID,useNamesName)

for m_id, score in topTrueFilmi:
    for mid, mname in skupajImena:
        if m_id==mid:
            movNames[mname]=score[0]
    skupajImena = zip(useNamesID, useNamesName)

print(movNames)
from openpyxl import *
topVrstica=[]
topVrstica.append("movieID")
for i in range(0,len(topTrueFilmi[0][1][0])+1):
    topVrstica.append("u" + str(i))
from openpyxl import *
wb = Workbook()
dest_filename = 'testing_data.xlsx'
ws1 = wb.active
ws1.title = "range names"
for ind,val in enumerate(topVrstica):
    temp = get_column_letter(ind+1)
    ws1[temp+str(1)].value=val

currentIndex=2
forKeys=1
for key,ratings in movNames.items():
    temp = get_column_letter(forKeys)
    ws1[temp + str(currentIndex)].value = key.strip("\"")
    for ind,rating in enumerate(ratings):
        ind=ind+1
        temp = get_column_letter(ind + 1)
        ws1[temp + str(currentIndex)].value = rating
    currentIndex+=1

wb.save(filename=dest_filename)

#for i in topTrueFilmi:
#    print(i[1][1][0])
