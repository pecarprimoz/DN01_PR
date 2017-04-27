from datetime import datetime
from numpy import *
from collections import defaultdict
from openpyxl.utils import get_column_letter

set_printoptions(suppress=True)

data = loadtxt("../data/ratings.csv",delimiter=",",skiprows=1)

userID=[int(i) for i in data[:,0]]
movieID=[int(i) for i in data[:,1]]
ratingID=[float(i) for i in data[:,2]]
timestamp=[datetime.fromtimestamp(i).strftime('%Y-%m-%d')
           for i in data[:,3]]
relevantData=zip(movieID,ratingID)
#vse skupne ocene, št. pojavitev
allMovies=defaultdict(list)
for mID,rating in relevantData:
    if(mID not in allMovies):
        allMovies[mID].append([0.0,0])
    allMovies[mID][0][1]+=1
    allMovies[mID][0][0]+=rating
#skupni rating (allMovies[key][0][0])
#št. pojavitev (allMovies[key][0][1])


#težava ker je film ocenjen enkrat ali manjkrat, popravimo tako da vzamemo vse filme ki so bili ocenjeni več kot 15x
trueAvgRatings=defaultdict(float)
for key,value in allMovies.items():
    if (allMovies[key][0][1] != 0 and allMovies[key][0][1]>100):
        trueAvgRatings[key] = allMovies[key][0][1]
    else:
        trueAvgRatings[key] = -1
#top 100 najbolj ogledanih filmov

topTrueFilmi = sorted(trueAvgRatings.items(), key=lambda v: v[1], reverse=True)[:100]

print(topTrueFilmi)

user_data=defaultdict(int)
for i in userID:
    user_data[i]+=1

topUserRatings=sorted(user_data.items(), key=lambda v: v[1], reverse=True)[:100]

print(topUserRatings)
film_matrix=defaultdict(lambda : defaultdict(int))
for key, value in topTrueFilmi:
    for u, rat in topUserRatings:
        film_matrix[key][u]=0

print(film_matrix)
a=1

for mid, uid, rating in zip(movieID,userID,ratingID):
    for key,value in film_matrix.items():
        if mid==key and uid in value:
            film_matrix[key][uid]=rating

print(film_matrix)
from openpyxl import load_workbook
movNames=defaultdict(list)
movieNames= load_workbook(filename="../data/moviesRMK_V1.xlsx")
useNames = movieNames['movies']
useNamesID=[]
useNamesName=[]
for i in range(2,9127):
    wholeColumn=useNames['A'+str(i)].value.split(",")
    useNamesID.append(wholeColumn[0])
    useNamesName.append(wholeColumn[1])

useNamesID=[int(i) for i in useNamesID]
skupajImena=zip(useNamesID,useNamesName)

topVrstica=[]
topVrstica.append("movieID")
for key,value in film_matrix.items():
    for temp in value.items():
        topVrstica.append("u"+str(temp[0]))
    break
print(len(topVrstica))
print(topVrstica)
from openpyxl import *
wb = Workbook()
dest_filename = '.' \
                './data/testing_data_regress.xlsx'

mov_names=[]
for key,val in film_matrix.items():
    temp=zip(useNamesID,useNamesName)
    for fid,mname in temp:
        if key==fid:
            mov_names.append(mname)
            break
print(len(mov_names))
print(mov_names)
ws1 = wb.active
ws1.title = "range names"
for ind,val in enumerate(topVrstica):
    temp = get_column_letter(ind+1)
    ws1[temp+str(1)].value=val

for ind,val in enumerate(mov_names):
    ws1['A' + str(ind+2)].value = val

currentIndex=2
forKeys=1
for key,value in film_matrix.items():
    for ind,val in enumerate(value.items()):
        ind = ind + 1
        temp = get_column_letter(ind + 1)
        ws1[temp + str(currentIndex)].value=val[1]
    currentIndex += 1

wb.save(filename=dest_filename)