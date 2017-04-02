from datetime import datetime
from numpy import *
from collections import defaultdict
#ID IME FILMA RATING
#858 GODFATHER

set_printoptions(suppress=True)

data = loadtxt("../data/ratings.csv",delimiter=",",skiprows=1)

movieID=[int(i) for i in data[:,1]]
ratingID=[float(i) for i in data[:,2]]
counter=0

timestamp=[datetime.fromtimestamp(i).strftime('%Y-%m-%d')
           for i in data[:,3]]

timestamp.sort()
relevantData=zip(movieID,ratingID,timestamp)

godRatings=list()
for mID,rating, time in relevantData:
    if(mID==858):
        godRatings.append([rating,time])

#print(godRatings)
avgRatings=defaultdict(float)
counter=0
currentScoreSum=0

for rating, time in godRatings:
    counter+=1
    currentScoreSum+=rating
    if(counter>30 and counter<=150):
        avgRatings[time]=currentScoreSum/counter
#print(avgRatings)

padajoceGod = sorted(avgRatings.items(), key=lambda v: v[0], reverse=False)
print("Statistika za godfatherja")
print(padajoceGod)
#test narejen na godfatherju pokaže da ocena kvečjemu narašča
#poskusimo še na Shawshank Redemption, film ki ima več kot 250 ocen in je tudi najboljši film z oceno 4.487138

data = loadtxt("../data/ratings.csv",delimiter=",",skiprows=1)

movieID=[int(i) for i in data[:,1]]
ratingID=[float(i) for i in data[:,2]]
counter=0

timestamp=[datetime.fromtimestamp(i).strftime('%Y-%m-%d')
           for i in data[:,3]]

timestamp.sort()
relevantData=zip(movieID,ratingID,timestamp)
shawkRatings=list()
for mID,rating, time in relevantData:
    if(mID==318):
        shawkRatings.append([rating,time])

#print(godRatings)
avgRatings=defaultdict(float)
counter=0
currentScoreSum=0

for rating, time in shawkRatings:
    counter+=1
    currentScoreSum+=rating
    if(counter>30 and counter<=150):
        avgRatings[time]=currentScoreSum/counter
print("Statistika za Shawshank")
padajoceShawk = sorted(avgRatings.items(), key=lambda v: v[0], reverse=False)
print(padajoceShawk)
#tudi tukaj vidimo da ocena kvečjemu narašča z časom

#še primer slabega filma, home alone več kot 100 ratingov z ratingom 3.112403
data = loadtxt("../data/ratings.csv",delimiter=",",skiprows=1)

movieID=[int(i) for i in data[:,1]]
ratingID=[float(i) for i in data[:,2]]
counter=0

timestamp=[datetime.fromtimestamp(i).strftime('%Y-%m-%d')
           for i in data[:,3]]

timestamp.sort()
relevantData=zip(movieID,ratingID,timestamp)
aloneRatings=list()
for mID,rating, time in relevantData:
    if(mID==586):
        aloneRatings.append([rating,time])

#print(godRatings)
avgRatings=defaultdict(float)
counter=0
currentScoreSum=0

for rating, time in aloneRatings:
    counter+=1
    currentScoreSum+=rating
    if(counter>30 and counter<=150):
        avgRatings[time]=currentScoreSum/counter
print("Statistika za Home alone")
padajoceAlone = sorted(avgRatings.items(), key=lambda v: v[0], reverse=False)

print(padajoceAlone)
#za zelo slab filem bi pričakoval da povprečna ocena pada, vendar v tem primeru narašča

#Charlie's Angels (2000) 3977
data = loadtxt("../data/ratings.csv",delimiter=",",skiprows=1)

movieID=[int(i) for i in data[:,1]]
ratingID=[float(i) for i in data[:,2]]
counter=0

timestamp=[datetime.fromtimestamp(i).strftime('%Y-%m-%d')
           for i in data[:,3]]

timestamp.sort()
relevantData=zip(movieID,ratingID,timestamp)
angelRatings=list()
for mID,rating, time in relevantData:
    if(mID==3977):
        angelRatings.append([rating,time])

avgRatings=defaultdict(float)
counter=0
currentScoreSum=0

for rating, time in angelRatings:
    counter+=1
    currentScoreSum+=rating
    if(counter>30 and counter<=150):
        avgRatings[time]=currentScoreSum/counter

print("Statistika za Charlie's Angels")
padajoceAngel = sorted(avgRatings.items(), key=lambda v: v[0], reverse=False)
print(padajoceAngel)
from openpyxl import Workbook
wb = Workbook()
'''
    Če bo kdo to v življenju bral, v primeru da hočeš točno iste stvari bo ta skripta delala lepo, vendar moraš
    spremeniti dest_filename, ta je prirejen za mojo win mašino.
'''
dest_filename= "../data/test_book.xlsx"
ws=wb.active
ws.title="test"
counter=1
for datum,val in padajoceAngel:
    ws["A"+str(counter)] =datum
    ws["B"+str(counter)] = val
    counter+=1

counter=1
for sDatum, sVal in padajoceShawk:
    ws["D" + str(counter)] = sDatum
    ws["E" + str(counter)] = sVal
    counter += 1

wb.save(filename=dest_filename)
